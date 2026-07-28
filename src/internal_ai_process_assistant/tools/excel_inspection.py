"""Safe Excel inspection utilities for controlled project input files."""

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from internal_ai_process_assistant.tools.input_file_validation import validate_input_file


@dataclass(frozen=True)
class ExcelSheetInfo:
    """Structured metadata for one Excel worksheet."""

    name: str
    row_count: int
    column_count: int


@dataclass(frozen=True)
class ExcelInspectionResult:
    """Structured metadata returned by the Excel inspection tool."""

    filename: str
    sheet_count: int
    sheets: list[ExcelSheetInfo]


def inspect_excel(filename: str, project_root: Path | None = None) -> ExcelInspectionResult:
    """Inspect a controlled Excel input file and return workbook metadata."""
    validated_file = validate_input_file(filename=filename, project_root=project_root)

    if validated_file.extension != ".xlsx":
        raise ValueError("Excel inspection requires a .xlsx file")

    root = project_root or Path.cwd()
    workbook_path = root / validated_file.relative_path

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheets = [
            ExcelSheetInfo(
                name=sheet.title,
                row_count=sheet.max_row,
                column_count=sheet.max_column,
            )
            for sheet in workbook.worksheets
        ]
    finally:
        workbook.close()

    return ExcelInspectionResult(
        filename=validated_file.filename,
        sheet_count=len(sheets),
        sheets=sheets,
    )
